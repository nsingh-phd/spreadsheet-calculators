#!/usr/bin/env python3
"""
Mortgage Amortization Calculator Generator

Generates a template Excel (.xlsx) spreadsheet with formula-driven
mortgage amortization schedule supporting:
- Recurring extra payments (monthly/annual) with configurable start dates
- One-time extra payments
- Summary comparison (with vs without extra payments)

Usage:
    python generate_mortgage_calculator.py

Output:
    mortgage_amortization_calculator.xlsx
"""

import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import CellIsRule
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.chart import PieChart, LineChart, Reference
from openpyxl.chart.series import DataPoint
from openpyxl.chart.label import DataLabelList
from copy import copy


def add_named_range(wb, name, attr_text):
    """Add a named range compatible with openpyxl 3.1.x."""
    dn = DefinedName(name, attr_text=attr_text)
    wb.defined_names.add(dn)

# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------
MAX_MONTHS = 360  # 30-year max
NUM_RECURRING = 10
NUM_ONETIME = 15

# Left panel layout rows (1-indexed)
INPUT_START_ROW = 2
SUMMARY_HEADER_ROW = 8
SUMMARY_DATA_START = 9
SUMMARY_DATA_END = 17

RECURRING_SECTION_ROW = 19
RECURRING_HEADER_ROW = 20
RECURRING_DATA_START = 21
RECURRING_DATA_END = RECURRING_DATA_START + NUM_RECURRING - 1  # 30

ONETIME_SECTION_ROW = 32
ONETIME_HEADER_ROW = 33
ONETIME_DATA_START = 34
ONETIME_DATA_END = ONETIME_DATA_START + NUM_ONETIME - 1  # 48

# Right panel: amortization schedule (starts at column H)
AMORT_COL = 8  # Column H
AMORT_HEADER_ROW = 2
AMORT_DATA_START = 3
AMORT_DATA_END = AMORT_DATA_START + MAX_MONTHS - 1  # 362

# Styling
HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
SECTION_FONT = Font(bold=True, size=12, color="2F5496")
INPUT_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
LIGHT_BLUE_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

# Currency and date formats
CURRENCY_FMT = '#,##0.00'
PERCENT_FMT = '0.000%'
DATE_FMT = 'MMM YYYY'
NUMBER_FMT = '#,##0'


def style_header_row_range(ws, row, col_start, col_end, fill=None, font=None):
    """Apply header styling to a range of cells in a row."""
    fill = fill or HEADER_FILL
    font = font or HEADER_FONT
    for c in range(col_start, col_end + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER


def style_input_cell(cell, fmt=None):
    """Style a cell as user-editable input."""
    cell.fill = INPUT_FILL
    cell.border = THIN_BORDER
    cell.alignment = Alignment(horizontal="center")
    if fmt:
        cell.number_format = fmt


def create_workbook():
    wb = Workbook()
    ws = wb.active
    ws.title = "Calculator"

    from datetime import date
    from openpyxl.worksheet.datavalidation import DataValidation

    # Column letters for amortization (starting at AMORT_COL=8 → H)
    AC = AMORT_COL
    ac = get_column_letter  # shorthand
    PAY_NUM_COL = AC       # H - Payment #
    DATE_COL = AC + 1      # I - Date
    PAYMENT_COL = AC + 2   # J - Payment
    INTEREST_COL = AC + 3  # K - Interest
    PRINCIPAL_COL = AC + 4 # L - Principal
    END_BAL_COL = AC + 5   # M - Ending Balance
    CUM_PRINCIPAL_COL = AC + 6  # N - Cumulative Principal (hidden)
    CUM_INTEREST_COL = AC + 7   # O - Cumulative Interest (hidden)
    NUM_AMORT_COLS = 6  # visible columns only

    # -----------------------------------------------------------------------
    # A. LOAN DETAILS (left panel, rows 1-6)
    # -----------------------------------------------------------------------
    ws.merge_cells("A1:B1")
    ws.cell(row=1, column=1, value="LOAN DETAILS").font = SECTION_FONT

    labels = [
        (2, "Loan Amount ($)"),
        (3, "Annual Interest Rate"),
        (4, "Loan Term (years)"),
        (5, "Start Date"),
        (6, "Monthly Payment"),
    ]
    for row_num, label in labels:
        ws.cell(row=row_num, column=1, value=label).font = Font(bold=True)
        ws.cell(row=row_num, column=1).alignment = Alignment(horizontal="right")

    style_input_cell(ws.cell(row=2, column=2, value=300000), CURRENCY_FMT)
    style_input_cell(ws.cell(row=3, column=2, value=0.05875), PERCENT_FMT)
    style_input_cell(ws.cell(row=4, column=2, value=30), NUMBER_FMT)
    style_input_cell(ws.cell(row=5, column=2), DATE_FMT)
    ws.cell(row=5, column=2).value = date(2026, 1, 1)

    ws.cell(row=6, column=2).value = '=ROUND(-PMT(B3/12,B4*12,B2),2)'
    ws.cell(row=6, column=2).number_format = CURRENCY_FMT
    ws.cell(row=6, column=2).font = Font(bold=True, color="006600")
    ws.cell(row=6, column=2).border = THIN_BORDER

    add_named_range(wb, "LoanAmount", "Calculator!$B$2")
    add_named_range(wb, "AnnualRate", "Calculator!$B$3")
    add_named_range(wb, "TermYears", "Calculator!$B$4")
    add_named_range(wb, "StartDate", "Calculator!$B$5")
    add_named_range(wb, "MonthlyPmt", "Calculator!$B$6")

    # -----------------------------------------------------------------------
    # B. LOAN SUMMARY (left panel, rows 8-17)
    # -----------------------------------------------------------------------
    ws.merge_cells("A8:B8")
    ws.cell(row=SUMMARY_HEADER_ROW, column=1, value="LOAN SUMMARY").font = SECTION_FONT

    # Amort column letter references for summary formulas
    pn = ac(PAY_NUM_COL)   # H
    dt = ac(DATE_COL)       # I
    pm = ac(PAYMENT_COL)    # J
    it = ac(INTEREST_COL)   # K
    pr = ac(PRINCIPAL_COL)  # L
    eb = ac(END_BAL_COL)    # M

    summary_items = [
        (9,  "Original Payoff Date"),
        (10, "Accelerated Payoff Date"),
        (11, "Accelerated Payoff Time"),
        (12, "Original Total Interest"),
        (13, "Accelerated Total Interest"),
        (14, "Interest Saved"),
        (15, "Months Saved"),
        (16, "Total of Payments (P+I)"),
        (17, "Total Extra Payments"),
    ]
    for row_num, label in summary_items:
        cell = ws.cell(row=row_num, column=1, value=label)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="right")

    vc = 2  # value column B

    ws.cell(row=9, column=vc, value='=EDATE(StartDate,TermYears*12-1)')
    ws.cell(row=9, column=vc).number_format = DATE_FMT

    ws.cell(row=10, column=vc,
            value=f'=LOOKUP(2,1/({dt}{AMORT_DATA_START}:{dt}{AMORT_DATA_END}<>""),{dt}{AMORT_DATA_START}:{dt}{AMORT_DATA_END})')
    ws.cell(row=10, column=vc).number_format = DATE_FMT

    actual_months_formula = f'SUMPRODUCT(1*(ISNUMBER({pn}{AMORT_DATA_START}:{pn}{AMORT_DATA_END})))'
    ws.cell(row=11, column=vc,
            value=f'=INT({actual_months_formula}/12)&" years, "&MOD({actual_months_formula},12)&" months"')

    ws.cell(row=12, column=vc, value='=ROUND(MonthlyPmt*TermYears*12-LoanAmount,2)')
    ws.cell(row=12, column=vc).number_format = CURRENCY_FMT

    ws.cell(row=13, column=vc,
            value=f'=ROUND(SUM({it}{AMORT_DATA_START}:{it}{AMORT_DATA_END}),2)')
    ws.cell(row=13, column=vc).number_format = CURRENCY_FMT

    ws.cell(row=14, column=vc, value=f'=B12-B13')
    ws.cell(row=14, column=vc).number_format = CURRENCY_FMT
    ws.cell(row=14, column=vc).font = Font(bold=True, color="006600", size=12)

    ws.cell(row=15, column=vc,
            value=f'=TermYears*12-SUMPRODUCT(1*(ISNUMBER({pn}{AMORT_DATA_START}:{pn}{AMORT_DATA_END})))')
    ws.cell(row=15, column=vc).number_format = NUMBER_FMT
    ws.cell(row=15, column=vc).font = Font(bold=True, color="006600", size=12)

    ws.cell(row=16, column=vc,
            value=f'=ROUND(SUM({pm}{AMORT_DATA_START}:{pm}{AMORT_DATA_END}),2)')
    ws.cell(row=16, column=vc).number_format = CURRENCY_FMT

    ws.cell(row=17, column=vc,
            value=f'=ROUND(SUM({pr}{AMORT_DATA_START}:{pr}{AMORT_DATA_END})-SUM({pm}{AMORT_DATA_START}:{pm}{AMORT_DATA_END})+SUM({it}{AMORT_DATA_START}:{it}{AMORT_DATA_END}),2)')
    ws.cell(row=17, column=vc).number_format = CURRENCY_FMT

    for row_num in range(9, 18):
        cell = ws.cell(row=row_num, column=vc)
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal="center")

    # -----------------------------------------------------------------------
    # C. RECURRING EXTRA PAYMENTS (left panel, rows 19-30)
    # -----------------------------------------------------------------------
    ws.merge_cells(f"A{RECURRING_SECTION_ROW}:E{RECURRING_SECTION_ROW}")
    ws.cell(row=RECURRING_SECTION_ROW, column=1, value="RECURRING EXTRA PAYMENTS").font = SECTION_FONT

    rec_headers = ["#", "Start Date", "End Date (optional)", "Amount ($)", "Frequency"]
    for i, h in enumerate(rec_headers, 1):
        ws.cell(row=RECURRING_HEADER_ROW, column=i, value=h)
    style_header_row_range(ws, RECURRING_HEADER_ROW, 1, 5)

    for idx in range(NUM_RECURRING):
        r = RECURRING_DATA_START + idx
        ws.cell(row=r, column=1, value=idx + 1).alignment = Alignment(horizontal="center")
        ws.cell(row=r, column=1).border = THIN_BORDER
        style_input_cell(ws.cell(row=r, column=2), DATE_FMT)       # Start Date
        style_input_cell(ws.cell(row=r, column=3), DATE_FMT)       # End Date (optional)
        style_input_cell(ws.cell(row=r, column=4), CURRENCY_FMT)   # Amount
        style_input_cell(ws.cell(row=r, column=5))                  # Frequency

    freq_dv = DataValidation(
        type="list",
        formula1='"Monthly,Annual"',
        allow_blank=True,
        showErrorMessage=True,
        errorTitle="Invalid Frequency",
        error="Please select Monthly or Annual",
    )
    freq_dv.add(f"E{RECURRING_DATA_START}:E{RECURRING_DATA_END}")
    ws.add_data_validation(freq_dv)

    # -----------------------------------------------------------------------
    # D. ONE-TIME EXTRA PAYMENTS (left panel, rows 32-48)
    # -----------------------------------------------------------------------
    ws.merge_cells(f"A{ONETIME_SECTION_ROW}:C{ONETIME_SECTION_ROW}")
    ws.cell(row=ONETIME_SECTION_ROW, column=1, value="ONE-TIME EXTRA PAYMENTS").font = SECTION_FONT

    ot_headers = ["#", "Date", "Amount ($)"]
    for i, h in enumerate(ot_headers, 1):
        ws.cell(row=ONETIME_HEADER_ROW, column=i, value=h)
    style_header_row_range(ws, ONETIME_HEADER_ROW, 1, 3)

    for idx in range(NUM_ONETIME):
        r = ONETIME_DATA_START + idx
        ws.cell(row=r, column=1, value=idx + 1).alignment = Alignment(horizontal="center")
        ws.cell(row=r, column=1).border = THIN_BORDER
        style_input_cell(ws.cell(row=r, column=2), DATE_FMT)
        style_input_cell(ws.cell(row=r, column=3), CURRENCY_FMT)

    # -----------------------------------------------------------------------
    # E. AMORTIZATION SCHEDULE (right panel, col H-M, rows 1-362)
    # -----------------------------------------------------------------------
    ws.merge_cells(f"{ac(AC)}1:{ac(AC + NUM_AMORT_COLS - 1)}1")
    ws.cell(row=1, column=AC, value="AMORTIZATION SCHEDULE").font = SECTION_FONT

    amort_headers = ["Payment #", "Date", "Payment", "Interest", "Principal", "Ending Balance"]
    for i, h in enumerate(amort_headers):
        ws.cell(row=AMORT_HEADER_ROW, column=AC + i, value=h)
    style_header_row_range(ws, AMORT_HEADER_ROW, AC, AC + NUM_AMORT_COLS - 1)

    # Hidden cumulative columns for chart (headers only for Reference titles_from_data)
    cp = ac(CUM_PRINCIPAL_COL)  # N
    ci = ac(CUM_INTEREST_COL)   # O
    ws.cell(row=AMORT_HEADER_ROW, column=CUM_PRINCIPAL_COL, value="Cumul. Principal")
    ws.cell(row=AMORT_HEADER_ROW, column=CUM_INTEREST_COL, value="Cumul. Interest")

    # Recurring/one-time range references (absolute)
    rec_start_range = f"$B${RECURRING_DATA_START}:$B${RECURRING_DATA_END}"
    rec_end_range = f"$C${RECURRING_DATA_START}:$C${RECURRING_DATA_END}"
    rec_amount_range = f"$D${RECURRING_DATA_START}:$D${RECURRING_DATA_END}"
    rec_freq_range = f"$E${RECURRING_DATA_START}:$E${RECURRING_DATA_END}"
    ot_date_range = f"$B${ONETIME_DATA_START}:$B${ONETIME_DATA_END}"
    ot_amount_range = f"$C${ONETIME_DATA_START}:$C${ONETIME_DATA_END}"

    for idx in range(MAX_MONTHS):
        r = AMORT_DATA_START + idx
        n = idx + 1
        prev_r = r - 1

        # Column letter references
        pn_cell = f"{pn}{r}"     # Payment #
        dt_cell = f"{dt}{r}"     # Date
        pm_cell = f"{pm}{r}"     # Payment
        it_cell = f"{it}{r}"     # Interest
        pr_cell = f"{pr}{r}"     # Principal
        eb_cell = f"{eb}{r}"     # Ending Balance

        # Beginning balance reference
        if idx == 0:
            bal_ref = "LoanAmount"
        else:
            bal_ref = f"{eb}{prev_r}"

        # Payment # — show if balance remaining
        if idx == 0:
            ws.cell(row=r, column=PAY_NUM_COL, value=f'=IF(LoanAmount>0,{n},"")')
        else:
            ws.cell(row=r, column=PAY_NUM_COL,
                    value=f'=IF(AND({eb}{prev_r}>0.005,{pn}{prev_r}<>""),{n},"")')

        # Date
        if idx == 0:
            ws.cell(row=r, column=DATE_COL, value=f'=IF({pn_cell}<>"",StartDate,"")')
        else:
            ws.cell(row=r, column=DATE_COL, value=f'=IF({pn_cell}<>"",EDATE(StartDate,{n - 1}),"")')
        ws.cell(row=r, column=DATE_COL).number_format = DATE_FMT

        # Interest
        ws.cell(row=r, column=INTEREST_COL,
                value=f'=IF({pn_cell}<>"",ROUND({bal_ref}*AnnualRate/12,2),"")')
        ws.cell(row=r, column=INTEREST_COL).number_format = CURRENCY_FMT

        # Payment (min of monthly pmt and balance + interest)
        ws.cell(row=r, column=PAYMENT_COL,
                value=f'=IF({pn_cell}<>"",MIN(MonthlyPmt,{bal_ref}+{it_cell}),"")')
        ws.cell(row=r, column=PAYMENT_COL).number_format = CURRENCY_FMT

        # Principal = scheduled principal + extra, capped at balance
        # End date check: (end="" OR date<=end) → ((rec_end="")+(rec_end<>"")*(rec_end>=date))
        end_date_check = f'(({rec_end_range}="")+({rec_end_range}<>"")*({rec_end_range}>={dt_cell}))'

        extra_sum = (
            # Recurring monthly (with end date check)
            f'SUMPRODUCT(({rec_start_range}<>"")*({rec_start_range}<={dt_cell})*{end_date_check}*({rec_freq_range}="Monthly")*{rec_amount_range})'
            f'+'
            # Recurring annual (month match + end date check)
            f'SUMPRODUCT(({rec_start_range}<>"")*({rec_start_range}<={dt_cell})*{end_date_check}*({rec_freq_range}="Annual")*(MONTH({rec_start_range})=MONTH({dt_cell}))*{rec_amount_range})'
            f'+'
            # One-time (year and month match)
            f'SUMPRODUCT(({ot_date_range}<>"")*(YEAR({ot_date_range})=YEAR({dt_cell}))*(MONTH({ot_date_range})=MONTH({dt_cell}))*{ot_amount_range})'
        )

        principal_formula = (
            f'=IF({pn_cell}<>"",'
            f'MIN({bal_ref},({pm_cell}-{it_cell})+{extra_sum}),'
            f'"")'
        )
        ws.cell(row=r, column=PRINCIPAL_COL, value=principal_formula)
        ws.cell(row=r, column=PRINCIPAL_COL).number_format = CURRENCY_FMT

        # Ending Balance
        ws.cell(row=r, column=END_BAL_COL,
                value=f'=IF({pn_cell}<>"",ROUND({bal_ref}-{pr_cell},2),"")')
        ws.cell(row=r, column=END_BAL_COL).number_format = CURRENCY_FMT

        # Cumulative Principal (hidden col for chart)
        if idx == 0:
            ws.cell(row=r, column=CUM_PRINCIPAL_COL,
                    value=f'=IF({pn_cell}<>"",{pr_cell},"")')
        else:
            ws.cell(row=r, column=CUM_PRINCIPAL_COL,
                    value=f'=IF({pn_cell}<>"",{cp}{prev_r}+{pr_cell},"")')
        ws.cell(row=r, column=CUM_PRINCIPAL_COL).number_format = CURRENCY_FMT

        # Cumulative Interest (hidden col for chart)
        if idx == 0:
            ws.cell(row=r, column=CUM_INTEREST_COL,
                    value=f'=IF({pn_cell}<>"",{it_cell},"")')
        else:
            ws.cell(row=r, column=CUM_INTEREST_COL,
                    value=f'=IF({pn_cell}<>"",{ci}{prev_r}+{it_cell},"")')
        ws.cell(row=r, column=CUM_INTEREST_COL).number_format = CURRENCY_FMT

        # Alternate row shading
        if idx % 2 == 1:
            for c in range(AC, AC + NUM_AMORT_COLS):
                ws.cell(row=r, column=c).fill = PatternFill(
                    start_color="F2F7FB", end_color="F2F7FB", fill_type="solid")

        # Borders and alignment
        for c in range(AC, AC + NUM_AMORT_COLS):
            ws.cell(row=r, column=c).border = THIN_BORDER
            ws.cell(row=r, column=c).alignment = Alignment(horizontal="center")

    # -----------------------------------------------------------------------
    # F. FORMATTING & FINAL TOUCHES
    # -----------------------------------------------------------------------
    col_widths = {
        1: 26,   # A - labels
        2: 18,   # B - values / start dates
        3: 20,   # C - end dates
        4: 16,   # D - amounts
        5: 14,   # E - frequency
        6: 3,    # F - spacer
        7: 3,    # G - spacer
        8: 12,   # H - payment #
        9: 14,   # I - date
        10: 16,  # J - payment
        11: 16,  # K - interest
        12: 16,  # L - principal
        13: 18,  # M - ending balance
    }
    for col, width in col_widths.items():
        ws.column_dimensions[get_column_letter(col)].width = width

    # Hide cumulative columns (used only by chart)
    ws.column_dimensions[ac(CUM_PRINCIPAL_COL)].width = 0.5
    ws.column_dimensions[ac(CUM_INTEREST_COL)].width = 0.5

    # Freeze panes: keep amort headers visible when scrolling
    ws.freeze_panes = f"{ac(AC)}{AMORT_DATA_START}"

    # -----------------------------------------------------------------------
    # G. CHARTS (far right, starting at column O)
    # -----------------------------------------------------------------------
    CHART_COL = CUM_INTEREST_COL + 3  # skip 2 cols for helper data

    # ggplot2-inspired soft colors
    SOFT_BLUE = "619CFF"    # principal
    SOFT_RED = "F8766D"     # interest
    SOFT_GREEN = "00BA38"   # balance

    # -- Pie chart helper data (hidden area) --
    HELPER_COL = CUM_INTEREST_COL + 1
    hc = ac(HELPER_COL)
    ws.cell(row=1, column=HELPER_COL, value="Category").font = Font(color="FFFFFF", size=1)
    ws.cell(row=2, column=HELPER_COL, value="Principal").font = Font(color="FFFFFF", size=1)
    ws.cell(row=3, column=HELPER_COL, value="Interest").font = Font(color="FFFFFF", size=1)
    ws.cell(row=1, column=HELPER_COL + 1, value="Amount").font = Font(color="FFFFFF", size=1)
    ws.cell(row=2, column=HELPER_COL + 1, value="=LoanAmount")
    ws.cell(row=2, column=HELPER_COL + 1).font = Font(color="FFFFFF", size=1)
    ws.cell(row=3, column=HELPER_COL + 1, value="=B13")
    ws.cell(row=3, column=HELPER_COL + 1).font = Font(color="FFFFFF", size=1)
    ws.column_dimensions[hc].width = 0.5
    ws.column_dimensions[ac(HELPER_COL + 1)].width = 0.5

    # -- PIE CHART: Principal vs Interest --
    pie = PieChart()
    pie.title = "Principal vs Interest"
    pie.style = 10
    pie.width = 10.5
    pie.height = 8.25

    labels_ref = Reference(ws, min_col=HELPER_COL, min_row=2, max_row=3)
    data = Reference(ws, min_col=HELPER_COL + 1, min_row=1, max_row=3)
    pie.add_data(data, titles_from_data=True)
    pie.set_categories(labels_ref)

    # Soft colors
    pt_principal = DataPoint(idx=0)
    pt_principal.graphicalProperties.solidFill = SOFT_BLUE
    pie.series[0].data_points.append(pt_principal)
    pt_interest = DataPoint(idx=1)
    pt_interest.graphicalProperties.solidFill = SOFT_RED
    pie.series[0].data_points.append(pt_interest)

    # No text labels inside pie — legend only
    pie.series[0].dLbls = DataLabelList()
    pie.series[0].dLbls.showPercent = False
    pie.series[0].dLbls.showCatName = False
    pie.series[0].dLbls.showVal = False
    pie.series[0].dLbls.showSerName = False

    ws.add_chart(pie, f"{ac(CHART_COL + 1)}1")

    # -- LINE CHART: Principal, Interest, Balance over time --
    line = LineChart()
    line.title = "Loan Amortization Over Time"
    line.style = 10
    line.width = 16.5
    line.height = 9.75

    # No axis titles
    line.y_axis.title = None
    line.x_axis.title = None

    # Y-axis: dollar ticks
    line.y_axis.numFmt = '$#,##0,K'
    line.y_axis.number_format = '$#,##0,K'
    line.y_axis.delete = False

    # Categories: dates (col I)
    cats = Reference(ws, min_col=DATE_COL, min_row=AMORT_DATA_START,
                     max_row=AMORT_DATA_END)

    # Series data: cumulative principal, cumulative interest, ending balance
    principal_data = Reference(ws, min_col=CUM_PRINCIPAL_COL,
                               min_row=AMORT_HEADER_ROW, max_row=AMORT_DATA_END)
    interest_data = Reference(ws, min_col=CUM_INTEREST_COL,
                              min_row=AMORT_HEADER_ROW, max_row=AMORT_DATA_END)
    balance_data = Reference(ws, min_col=END_BAL_COL,
                             min_row=AMORT_HEADER_ROW, max_row=AMORT_DATA_END)

    line.add_data(principal_data, titles_from_data=True)
    line.add_data(interest_data, titles_from_data=True)
    line.add_data(balance_data, titles_from_data=True)
    line.set_categories(cats)

    # Soft colors matching pie + green for balance, wider lines
    line.series[0].graphicalProperties.line.solidFill = SOFT_BLUE
    line.series[1].graphicalProperties.line.solidFill = SOFT_RED
    line.series[2].graphicalProperties.line.solidFill = SOFT_GREEN

    for s in line.series:
        s.graphicalProperties.line.width = 28000  # ~2.2pt
        s.smooth = False

    # X-axis: ticks at 5-year intervals (60 months)
    line.x_axis.numFmt = 'YYYY'
    line.x_axis.majorTimeUnit = "years"
    line.x_axis.number_format = 'YYYY'
    line.x_axis.tickLblPos = "low"
    line.x_axis.delete = False
    line.x_axis.majorUnit = 60  # 5 years in months
    # Very faint gray gridlines
    from openpyxl.drawing.line import LineProperties
    from openpyxl.chart.axis import ChartLines
    from openpyxl.drawing.fill import GradientFillProperties
    faint_gray_lp = LineProperties(solidFill="E0E0E0", w=6350)
    from openpyxl.chart.shapes import GraphicalProperties
    faint_props = GraphicalProperties(ln=faint_gray_lp)
    line.y_axis.majorGridlines = ChartLines(spPr=faint_props)
    line.x_axis.majorGridlines = ChartLines(spPr=faint_props)

    # Place line chart below pie chart
    ws.add_chart(line, f"{ac(CHART_COL + 1)}14")

    return wb, ws


def main():
    wb, ws = create_workbook()

    # Save
    output_path = os.path.join(os.path.dirname(__file__), "mortgage_amortization_calculator.xlsx")
    wb.save(output_path)
    print(f"✅ Mortgage calculator generated: {output_path}")
    print(f"   Open in Excel and edit the yellow input cells.")
    print(f"   - Loan details: B2–B5")
    print(f"   - Recurring extra payments: rows {RECURRING_DATA_START}–{RECURRING_DATA_END} (cols B-E)")
    print(f"   - One-time extra payments: rows {ONETIME_DATA_START}–{ONETIME_DATA_END} (cols B-C)")
    print(f"   - Amortization schedule: columns H-M")


if __name__ == "__main__":
    main()
